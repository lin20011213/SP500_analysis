import requests
from bs4 import BeautifulSoup
import pdfplumber
import openpyxl
import os

def read_excel(name,corp_id):
    #分析表檔名
    EXCEL_NAME='Sustainability-template-Sean-Cheng-Bachelors-Thesis.xlsx'
    TARGET_FLODER='Individual_Sustainability_Report'
    folder_path = "word"
    file_list = os.listdir(folder_path)
    # 開啟 Excel 檔案檔名
    workbook = openpyxl.load_workbook(EXCEL_NAME)
    workbook2 = openpyxl.load_workbook(EXCEL_NAME)
    # 選擇要讀取的工作表
    worksheet = workbook['Sheet1']
    worksheet_new = workbook2['Sheet1']
    if 'Sheet2' in workbook.sheetnames:
        worksheet2 = workbook['Sheet2']
        del workbook2['Sheet2']
    else:
        workbook.create_sheet("Sheet2")
        worksheet2 = workbook['Sheet2']
    # 讀取單元格資料
    list_num=0
    for file_name in file_list:
        try:
            if 'txt' not in file_name:
                continue
            file_path = os.path.join(folder_path, file_name)
            list_num=list_num+1
            if 'NO_' in file_name[:-9]:
                id=file_name[3:-9]
                NO=True
            else:
                id=file_name[:-9]
                rate_name=file_name[-8:-4]
                NO=False
            if rate_name == '2023':
                rate=1.1
            elif rate_name == '2022':
                rate=1.0
            elif rate_name == '2021':
                rate=0.9
            elif rate_name == '2020':
                rate=0.8
            elif rate_name == '2019':
                rate=0.7
            else:
                rate=0.6
            #print(rate_name)
            print(id)
            corp_name = name[corp_id.index(id)]
            worksheet2[f'A{list_num}'].value=list_num
            worksheet2[f'B{list_num}'].value=corp_name
            worksheet2[f'C{list_num}'].value=id
            
            if os.path.isfile(file_path):
                with open(file_path, 'r',encoding="utf-8") as file:
                    content = file.read()
                    content=content.lower()
                    #print(content)
                    sum=0.00
                    target_value=0.00
                    #print(worksheet2.max_row)
                    #print(worksheet.max_column)
                    for num in range(2,worksheet.max_row):
                        cell_value = worksheet[f'B{num}'].value
                        if cell_value == None:
                            break
                        #print(num)
                        #print(cell_value)
                        count=content.count(cell_value.lower())
                        if count ==0:
                            score=0.00
                        elif count ==1:
                            score=0.33
                        elif count ==2:
                            score=0.66
                        elif count >2:
                            score=1.00
                        worksheet_new[f'E{num}'].value=str(score)
                        worksheet_new[f'F{num}'].value=str(score*worksheet[f'D{num}'].value)
                        sum=sum+score*worksheet[f'D{num}'].value
                        target_value=worksheet[f'D{num}'].value+target_value
                    worksheet2[f'D{list_num}'].value=str(sum*rate)
                    sus=sum*rate/target_value
                    #print(sus)
                    if sus >= 0.666:
                        Sustainability_rating='sustainable'
                    elif sus >= 0.333:
                        Sustainability_rating='netural'
                    else:
                        Sustainability_rating='Avoid'
                    if (NO):
                        worksheet2[f'E{list_num}'].value="No file available"
                    else:
                        worksheet2[f'E{list_num}'].value=Sustainability_rating
                    worksheet_new['G2'].value=sum
                    worksheet_new['H2'].value=Sustainability_rating
                    #print(sum)
                    workbook2.save(f'{TARGET_FLODER}/{id}.xlsx')
        except:
            continue
    # 關閉 Excel 檔案
    workbook.save('Sustainability-template-Sean-Cheng-Bachelors-Thesis.xlsx')
    workbook.close()
    workbook2.close()

def extract_text_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()
    return text

def get_corp_list():
    corp_list=[]
    corp_name=[]
    header={'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.57'}
    url="https://www.slickcharts.com/sp500"
    resp=requests.get(url,headers=header)
    soup = BeautifulSoup(resp.text, 'html.parser')
    thing=soup.find('div',class_='col-lg-7')
    list=thing.find('tbody').find_all('tr')
    for data in list:
        lista=data.find_all('td')
        corp_name.append(lista[1].get_text())
        corp_list.append(lista[2].get_text())
    return corp_name,corp_list

def get_and_save_pdf(list):
    for corp in list:
        try:
            year='2023'
            url=f"https://www.responsibilityreports.com/HostedData/ResponsibilityReports/PDF/NASDAQ_{corp}_{year}.pdf"
            resp=requests.get(url)
            while(resp.status_code == 404 or int(year)<2019):
                year=str(int(year)-1)
                if resp.status_code == 404:
                    url=f"https://www.responsibilityreports.com/HostedData/ResponsibilityReports/PDF/NYSE_{corp}_{year}.pdf"
                    resp=requests.get(url)
                if resp.status_code == 404:
                    url=f"https://www.responsibilityreports.com/HostedData/ResponsibilityReports/PDF/NASDAQ_{corp}_{year}.pdf"
                    resp=requests.get(url)
            with open(f"pdf/{corp}_{year}.pdf", 'wb') as file:
                    file.write(resp.content)
            resp=extract_text_from_pdf(f"pdf/{corp}_{year}.pdf")
            with open(f"word/{corp}_{year}.txt", 'wb') as file:
                    file.write(resp.encode())
        except:

            with open(f"word/NO_{corp}_{year}.txt", 'wb') as file:
                    file.write("NO file".encode())

if __name__ == '__main__':
    '''
    要記得刪掉 # 字號才可以運作，要使用哪種程式，刪那種前的井字號
    
    '''
    #分析 字跟讀取EXCEL
    name,corp_list=get_corp_list()
    read_excel(name,corp_list)